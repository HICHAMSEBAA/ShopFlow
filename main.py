#!/home/hicham/Hicham/Python/ShopFlow/VShopFlow/bin/python3

import tkinter as tk
from tkinter import ttk
import openpyxl
import os
from datetime import datetime
from tkinter import messagebox
import uuid
import threading
from tkcalendar import DateEntry


class ExcelCrafterApp:
    def __init__(self, root):
        # Initialize the main window (root) and set the title
        self.root = root
        self.root.title("Product Management")
        
        # Initialize the styles for the application
        self.style = ttk.Style(root)
        root.tk.call("source", "forest-light.tcl")  # Load the custom theme from a Tcl file
        self.style.theme_use("forest-light")        # Apply the "forest-light" theme
        
        # Initialize product-related dropdown options (ComboBox lists)
        self.combo_list_ProductCategory = ["Adapter", "Cable", "Car Accessory", "Car Charger", "Charger", "Flash USB", "Earphone", "Memory Card", "Shock Glasses", "SIM Card", "Other"] 
        self.combo_list_ProductType = ["Auxiliary", "Bluetooth", "iPhone", "Micro USB", "Simple Phone", "Smart Phone", "Type C", "Wired", "Other"]
        self.combo_list_MemoryType = ["1 GB", "2 GB", "4 GB", "8 GB", "16 GB", "32 GB", "64 GB", "128 GB", "256 GB"]
        self.combo_list_SIMCardType = ["GOLD", "DIMA","DIMA+", "YOOZ", "LEGEND", "LEGEND", "ZID", "MOBTASIM", "SAMA", "Normal", "Other"]
        
        # Initialize beverage-related
        self.combo_list_beverageCategory = ["Soft Drinks", "Juices", "Water", "Energy Drinks", "Tea & Coffee", "Alcoholic Beverages", "Dairy-Based Drinks", "Smoothies", "Mocktails", "Herbal and Health Drinks" ] 
        self.combo_list_beveragebrand = ["Coca-Cola", "Pepsi", "7UP", "Sprite", "Fanta"]

    # product section ---------------------------------------------------------------------
    
        # widgets for the product management section.
        self.product_name_entry = None
        self.category_combobox = None
        self.type_combobox = None
        self.memoryType_combobox = None
        self.quantity_spinbox = None
        self.price_spinbox = None
        self.update_button = None
        self.delete_button = None
        self.cancel_button = None
        self.Sales_button = None
        
        # widgets for the products sales management section.
        self.sales_name_entry  = None
        self.sales_category_combobox = None
        self.sales_type_entry  = None
        self.sales_quantity_spinbox = None
        self.sales_price_spinbox = None
        self.sales_return_button = None
        self.sales_cancel_button = None
        
    # product section ---------------------------------------------------------------------
    
    
    # Beverage section ---------------------------------------------------------------------
    
        # widgets for the beverages management 
        self.beverage_name_entry  = None
        self.beverage_category_combobox = None
        self.beverage_brand_combobox  = None
        self.beverage_quantity_spinbox = None
        self.beverage_price_spinbox = None
        self.beverage_save_button = None
        self.beverage_update_button = None
        self.beverage_delete_button = None
        self.beverage_cancel_button = None
        self.beverage_Sales_button = None
        

        
        # Create a single search entry attribute
        self.product_search_entry = None
        self.sales_search_entry = None
        self.beverage_search_entry = None
        
        
        # Initialize Tables View
        self.treeview1 = None
        self.treeview2 = None
        self.treeview3 = None
        
        # Data load initialization 
        self.data_product = None
        self.data_sales = None
        self.data_beverage = None
        
        # Initialize a variable to keep track of the selected item in the Treeview1
        self.selected_prodcut_item = None
        
        # Initialize a variable to keep track of the selected item in the Treeview1
        self.selected_beverage_item = None
        
        
        # Initialize a variable to keep track of the selected item in the Treeview2
        self.selected_sales_item = []
        
        # RESET Product Flage
        self.reset_product_flag = False
        
        # RESET Beveratge Flage
        self.reset_beverage_flag = False
        
        # RESET Sales Flage
        self.reset_sales_flag = False
        
        # RETURN Sales Flage
        self.return_sales_flag = False
        
        # Create the main frame that will contain all other widgets
        # Create the main frame inside the root window
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.grid(row=0, column=0, sticky="nsew")

        # Configure the grid of the root window to make the main frame expandable
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # Optionally, configure the grid inside the main frame if it contains child widgets
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        # Expand the frame to fit the window
        
        # Create a notebook (tabbed interface) within the main frame
        self.create_notebook(self.main_frame)

        # products ----------------------------------------------------------------
        # Add a title to the first tab (Product Page)
        self.page_title(self.tab1, "The Product Page")
        
        # Add product management widgets (input fields, buttons, etc.) to the first tab
        self.add_product_widgets(self.tab1)
        
        # Create a Treeview widget (a table-like structure) in the first tab to display product data
        self.create_treeview("treeview1", self.tab1, path="products.xlsx", columns=["Name", "Category", "Type", "Quantity", "Price", "Date", "Time"])
        
        # Create a search bar in the first tab to allow users to search through products
        self.create_search(self.tab1, "product_search_entry", path="products.xlsx")
        # products ----------------------------------------------------------------
        
        # sales products ----------------------------------------------------------------
        # Add a title to the first tab (Product Page)
        self.page_title(self.tab2, "The Sales Page")
        
        # Create a Treeview widget (a table-like structure) in the first tab to display product data
        self.create_treeview("treeview2", self.tab2, path="sales.xlsx", columns=["Name", "Category", "Type", "Quantity", "Price", "Date", "Time"])
        
        # Add product management widgets (input fields, buttons, etc.) to the first tab
        self.add_sales_widgets(self.tab2)
        
        # Create a search bar in the first tab to allow users to search through products
        self.create_search(self.tab2, "sales_search_entry", path="sales.xlsx")
        # sales -------------------------------------------------------------------------
        
        # beverage section ----------------------------------------------------------------
        # Add a title to the first tab (Product Page)
        self.page_title(self.tab3, "The Beverage Page")
        
        # Create a Treeview widget (a table-like structure) in the first tab to display product data
        self.create_treeview("treeview3", self.tab3, path="beverage.xlsx", columns=["Name", "Category", "Brand", "Quantity", "Price", "Date", "Time"])
        
        # Add product management widgets (input fields, buttons, etc.) to the first tab
        self.add_berevage_widgets(self.tab3)
        
        # Create a search bar in the first tab to allow users to search through products
        self.create_search(self.tab3, "beverage_search_entry", path="beverage.xlsx")
        
        # beverage section ----------------------------------------------------------------
        
        
    
# Sales products part ----------------------------------------------------------------------------

    def open_sales_window(self):
        """
        Opens a new window for processing a sale of the selected product.
        """
        # Get the selected item from the Treeview1
        selected_item = self.treeview1.selection()
        
        # If no item is selected, show an error message and exit the function
        if not selected_item:
            messagebox.showerror("Error", "Please select a product first.")
            return
        
        # Try to retrieve the details of the selected product
        try:
            product_details = self.treeview1.item(selected_item, "values")
            # Unpack the product details
            product_name, category, product_type, available_quantity, price, date, time = product_details
            
            # Check if the product is out of stock
            if int(available_quantity) <= 0:
                messagebox.showinfo("Info", "The product is out of stock.")
                return
        except Exception as e:
            # If there's an error retrieving product details, show an error message
            messagebox.showerror("Error", f"Failed to retrieve product details: {e}")
            return
        
        # Create a new top-level window for the sales operation
        sales_window = tk.Toplevel(self.root)
        sales_window.title("Sales Operation")
        
        # Determine which type of sales window to create based on the product category
        if category == "SIM Card":
            self.create_sales_window(
                sales_window, 
                product_name, 
                category, 
                product_type, 
                available_quantity, 
                price, 
                product_type_list=self.combo_list_SIMCardType, 
                window_title="Sales SIMCards"
            )
        else:
            self.create_sales_window(
                sales_window, 
                product_name, 
                category, 
                product_type, 
                available_quantity, 
                price, 
                product_type_list=self.combo_list_SIMCardType, 
                window_title="Sales Products"
            )

    def create_sales_window(self, sales_window, product_name, category, product_type, available_quantity, price, product_type_list, window_title):
        """
        Creates and configures the sales window with all necessary widgets.
        
        Parameters:
        - sales_window: The Toplevel window where the sales operation will take place.
        - product_name: Name of the product being sold.
        - category: Category of the product.
        - product_type: Type of the product.
        - available_quantity: Current stock quantity of the product.
        - price: Price of the product.
        - product_type_list: List of product types for the Combobox.
        - window_title: Title of the sales window.
        """
        # Create the main frame within the sales window
        main_frame = ttk.Frame(sales_window)
        main_frame.pack(fill="both", expand=True)

        # Create a labeled frame to hold the sales widgets
        frame_widgets = ttk.LabelFrame(main_frame, text=window_title)
        frame_widgets.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        
        # ------------------ Product Name ------------------
        # Label for product name
        name_label = ttk.Label(frame_widgets, text="Name:")
        name_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        # Entry widget for product name (disabled as it's pre-filled)
        product_name_entry = ttk.Entry(frame_widgets)
        product_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        product_name_entry.insert(0, product_name)  # Set the initial value
        product_name_entry.config(state="disabled")  # Disable the entry to prevent editing
        
        # ------------------ Category ------------------
        # Label for category
        category_label = ttk.Label(frame_widgets, text="Category:")
        category_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        
        # Combobox for category (disabled as it's pre-filled)
        category_combobox = ttk.Combobox(
            frame_widgets, 
            values=self.combo_list_ProductCategory, 
            state="readonly"
        )
        category_combobox.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        category_combobox.set(category)  # Set the initial value
        category_combobox.config(state="disabled")  # Disable editing
        
        # ------------------ Type ------------------
        # Label for product type
        type_label = ttk.Label(frame_widgets, text="Type:")
        type_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        
        # Combobox for product type
        type_combobox = ttk.Combobox(
            frame_widgets, 
            values=product_type_list, 
            state="readonly"
        )
        type_combobox.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        
        # Set the initial value and configure state based on window title
        if window_title != "Sales SIMCards":
            type_combobox.set(product_type)
            type_combobox.config(state="disabled")  # Disable if not SIM Cards
        else:
            type_combobox.set("Normal")  # Default value for SIM Cards
            type_combobox.config(state="normal")  # Enable editing
        
        # ------------------ Quantity ------------------
        # Label for quantity to sell
        quantity_label = ttk.Label(frame_widgets, text="Quantity:")
        quantity_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        
        # Spinbox for selecting quantity to sell (from 0 to available_quantity)
        quantity_spinbox = ttk.Spinbox(
            frame_widgets, 
            from_=0, 
            to=int(available_quantity)
        )
        quantity_spinbox.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
        
        # ------------------ Price ------------------
        # Label for price
        price_label = ttk.Label(frame_widgets, text="Price:")
        price_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
        
        # Spinbox for setting the price (from 0.0 to 10000.0 with increments of 50)
        price_spinbox = ttk.Spinbox(
            frame_widgets, 
            from_=0.0, 
            to=10000.0, 
            increment=50
        )
        price_spinbox.grid(row=4, column=1, padx=5, pady=5, sticky="ew")
        
        # Set the initial price value and configure state based on window title
        if window_title != "Sales SIMCards":
            price_spinbox.insert(0, price)  # Set the initial price
            price_spinbox.config(state="disabled")  # Disable if not SIM Cards
        else:
            price_spinbox.insert(0, 0)  # Default price for SIM Cards
            price_spinbox.config(state="normal")  # Enable editing
        
        # ------------------ Sell Button ------------------
        # Button to execute the sell operation
        sell_button = ttk.Button(
            frame_widgets, 
            text="Sell", 
            command=lambda: self.sell_product_from_window(
                product_name, 
                category_combobox, 
                type_combobox, 
                quantity_spinbox,
                price_spinbox,
                sales_window
            )
        )
        sell_button.grid(row=6, column=0, columnspan=2, padx=5, pady=10, sticky="nsew")

    def sell_product_from_window(self, product_name, category_combobox, type_combobox, quantity_spinbox, price, window):
        """
        Processes the sale of a product by updating the inventory and recording the sale.
        
        Parameters:
        - product_name: Name of the product being sold.
        - category_combobox: Combobox widget for product category.
        - type_combobox: Combobox widget for product type.
        - quantity_spinbox: Spinbox widget for quantity to sell.
        - window: The sales window to be closed after processing the sale.
        """
        # Validate the quantity entered by the user
        try:
            quantity_sold = int(quantity_spinbox.get())  # Get the quantity to sell
            if quantity_sold <= 0:
                # If quantity is not positive, show an error message
                messagebox.showerror("Error", "Please enter a valid quantity.")
                return
        except Exception as e:
            # If there's an error parsing the quantity, show an error message
            messagebox.showerror("Error", f"An error occurred while processing the quantity: {e}")
            return
        
        # Show confirmation dialog
        response = messagebox.askyesno("Confirm Sell", "Are you sure you want to Sell this product?")
        
        if  response:
            
            # Path to the products Excel file
            path = "products.xlsx"
            
            try:
                # Load the Excel workbook and select the active sheet
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active

                # Iterate through the rows to find the matching product
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    if row[0].value == product_name:
                        available_quantity = int(row[3].value)  # Current stock
                        
                        # Check if there is sufficient stock to fulfill the sale
                        if quantity_sold > available_quantity:
                            # If not enough stock, show a warning message
                            messagebox.showwarning("Insufficient Stock", f"Only {available_quantity} units available.")
                            return
                        else:
                            # Deduct the sold quantity from available stock
                            row[3].value = available_quantity - quantity_sold
                            # Prepare the updated data for the Treeview
                            new_data = [row[i].value for i in range(len(row))]
                            break
                else:
                    # If the product was not found in the Excel sheet, show a warning
                    messagebox.showwarning("Product Not Found", "The specified product was not found.")
                    return

                # Save the updated workbook back to the file
                workbook.save(path)
                
                # Record the sale in the sales log
                self.record_sale(
                    product_name, 
                    str(category_combobox.get()), 
                    price.get(),
                    str(type_combobox.get()), 
                    quantity_sold
                )
                
                # Close the sales window after successful sale
                window.destroy()

                # Notify the user of the successful sale
                messagebox.showinfo("Success", "Product sold successfully!")
                
                # Update the Treeview1 to reflect the new quantity
                for item in self.treeview1.get_children():
                    if self.treeview1.item(item, "values")[0] == product_name:
                        self.treeview1.item(item, values=new_data)
                        break
                
                # Reset any necessary variables or states
                self.reset_product()
                
            except Exception as e:
                # If any error occurs during the sale process, show an error message
                messagebox.showerror("Error", f"An error occurred while selling the product: {e}")

    def record_sale(self, product_name, category, price, type, quantity_sold):
        """
        Records the sale details into the sales Excel file.
        
        Parameters:
        - product_name: Name of the product sold.
        - category: Category of the product.
        - type: Type of the product.
        - quantity_sold: Quantity of the product sold.
        """
        # Path to the sales Excel file
        sales_path = "sales.xlsx"
        
        # Get the current date and time for the sale record
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")

        # Generate a random UUID
        unique_id = uuid.uuid4()

        # Convert to string if needed
        unique_id_str = str(unique_id)
    
        # Create a list representing the sale record
        sale_record = [unique_id_str, product_name, category, type, quantity_sold, price, current_date, current_time]

        try:
            if not os.path.exists(sales_path):
                # If the sales file doesn't exist, create it and add headers
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                headers = ["Id", "Product Name", "Category", "Type", "Quantity Sold", "price", "Date", "Time"]
                sheet.append(headers)
            else:
                # If the sales file exists, load it
                workbook = openpyxl.load_workbook(sales_path)
                sheet = workbook.active

            # Append the sale record to the sales sheet
            sheet.append(sale_record)
            
            # Save the workbook
            workbook.save(sales_path)
            
            # Insert into Treeview
            self.treeview2.insert('', tk.END, values=sale_record)
            
        except Exception as e:
            # If there's an error recording the sale, show an error message
            messagebox.showerror("Error", f"An error occurred while recording the sale: {e}")

# ---------------------------------------------------------------------------------------

# GUI MAIN PART ------------------------------------------------------------------------------

    def create_notebook(self, frame):
        """Creates the notebook and adds tabs to it."""
        # Create the notebook widget
        self.notebook = ttk.Notebook(frame)
        self.notebook.grid(row=0, column=0, sticky="nsew")

        # Configure grid weights for the notebook to make it responsive
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        
        # Create the first tab (Products)
        self.tab1 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab1, text="Products")
        
        # Configure grid weights for the tab1 to make it expandable
        self.tab1.grid_rowconfigure(1, weight=1)
        self.tab1.grid_columnconfigure(1, weight=1)
        
        # Create the second tab (Sales)
        self.tab2 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab2, text="Sales")
        
        # Configure grid weights for the tab2 to make it expandable
        self.tab2.grid_rowconfigure(1, weight=1)
        self.tab2.grid_columnconfigure(1, weight=1)
        
        # Create the second tab (Sales)
        self.tab3 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab3, text="Beverege")
        
        # Configure grid weights for the tab2 to make it expandable
        self.tab3.grid_rowconfigure(1, weight=1)
        self.tab3.grid_columnconfigure(1, weight=1)
        
        # Bind the <<NotebookTabChanged>> event
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_selected)
        
    def on_tab_selected(self, event):
        selected_tab = self.notebook.index(self.notebook.select())  # Get the index of the selected tab

        if selected_tab == 0:  # If Tab 1 is selected
            self.reset_product_cancel()
        elif selected_tab == 1:  # If Tab 2 is selected
            self.reset_product()

    def add_product_widgets(self, frame):
        """Creates and adds the widgets for the product management section."""
        frame_widgets = ttk.LabelFrame(frame, text="Products")
        frame_widgets.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        
        # Configure grid weights to make columns and rows responsive
        frame_widgets.grid_columnconfigure(1, weight=1)  # Make the second column (with entry fields) expandable
        for i in range(10):
            frame_widgets.grid_rowconfigure(i, weight=1)  # Make each row expandable

        
        
        # Product Name
        name_label = ttk.Label(frame_widgets, text="Name:")
        name_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.product_name_entry = ttk.Entry(frame_widgets)
        self.product_name_entry.bind("<FocusIn>", lambda e: self.clear_entry(self.product_name_entry, "Name"))
        self.product_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        # Category
        category_label = ttk.Label(frame_widgets, text="Category:")
        category_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.category_combobox = ttk.Combobox(frame_widgets, values=self.combo_list_ProductCategory, state="readonly")
        self.category_combobox.current(0)  # Default to the first category
        self.category_combobox.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.category_combobox.bind("<<ComboboxSelected>>", self.on_category_selected)  # Bind category selection event

        # Type
        type_label = ttk.Label(frame_widgets, text="Type:")
        type_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.type_combobox = ttk.Combobox(frame_widgets, values=self.combo_list_ProductType, state="readonly")
        self.type_combobox.current(0)
        self.type_combobox.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        self.type_combobox.config(state="disabled")

        # Memory Type
        type_label_ = ttk.Label(frame_widgets, text="Memory Type:")
        type_label_.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.memoryType_combobox = ttk.Combobox(frame_widgets, values=self.combo_list_MemoryType, state="readonly")
        self.memoryType_combobox.current(0)
        self.memoryType_combobox.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
        self.memoryType_combobox.config(state="disabled")

        # Quantity
        quantity_label = ttk.Label(frame_widgets, text="Quantity:")
        quantity_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.quantity_spinbox = ttk.Spinbox(frame_widgets, from_=1, to=1000)
        self.quantity_spinbox.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

        # Price
        price_label = ttk.Label(frame_widgets, text="Price:")
        price_label.grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.price_spinbox = ttk.Spinbox(frame_widgets, from_=0.0, to=10000.0, increment=50)
        self.price_spinbox.grid(row=5, column=1, padx=5, pady=5, sticky="ew")

        # Buttons for product operations
        self.insert_button = ttk.Button(frame_widgets, text="Insert", command=self.insert_product)
        self.insert_button.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")

        self.update_button = ttk.Button(frame_widgets, text="Update", command= lambda :self.update_product(treeview=self.treeview1))
        self.update_button.grid(row=6, column=1, padx=5, pady=5, sticky="nsew")
        self.update_button.config(state="disabled")

        self.delete_button = ttk.Button(frame_widgets, text="Delete", command=lambda: self.delete_product(treeview=self.treeview1))
        self.delete_button.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")
        self.delete_button.config(state="disabled")

        self.cancel_button = ttk.Button(frame_widgets, text="Cancel", command=self.reset_product)
        self.cancel_button.grid(row=7, column=1, padx=5, pady=5, sticky="nsew")
        self.cancel_button.config(state="disabled")

        # Separator for better UI structure
        separator = ttk.Separator(frame_widgets)
        separator.grid(row=8, column=0, columnspan=2, padx=(20, 10), pady=10, sticky="ew")
        
        # Sales Button
        self.Sales_button = ttk.Button(frame_widgets, text="Sale", command=self.open_sales_window)
        self.Sales_button.grid(row=9, column=0, columnspan=2, padx=5, pady=10, sticky="nsew")
        self.Sales_button.config(state="disable")
       
    def add_sales_widgets(self, frame):
        """Creates and adds the widgets for the product management section."""
        frame_widgets = ttk.LabelFrame(frame, text="Sales")
        frame_widgets.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        
        # Configure grid weights to make columns and rows responsive
        frame_widgets.grid_columnconfigure(1, weight=1)  # Make the second column (with entry fields) expandable
        for i in range(9):
            frame_widgets.grid_rowconfigure(i, weight=1)  # Make each row expandable
            
        # Product Name
        name_label = ttk.Label(frame_widgets, text="Name:")
        name_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.sales_name_entry = ttk.Entry(frame_widgets)
        self.sales_name_entry.bind("<FocusIn>", lambda e: self.clear_entry(self.product_name_entry, "Name"))
        self.sales_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.sales_name_entry.config(state="disabled")

        # Category
        category_label = ttk.Label(frame_widgets, text="Category:")
        category_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.sales_category_combobox = ttk.Combobox(frame_widgets, values=self.combo_list_ProductCategory, state="readonly")
        self.sales_category_combobox.current(0)  # Default to the first category
        self.sales_category_combobox.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.sales_category_combobox.bind("<<ComboboxSelected>>", self.on_category_selected)  # Bind category selection event
        self.sales_category_combobox.config(state="disabled")

        # Type
        type_label = ttk.Label(frame_widgets, text="Type:")
        type_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.sales_type_entry = ttk.Entry(frame_widgets)
        self.sales_type_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        self.sales_type_entry.config(state="disabled")

        # Quantity
        quantity_label = ttk.Label(frame_widgets, text="Quantity:")
        quantity_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.sales_quantity_spinbox = ttk.Spinbox(frame_widgets, from_=1, to=1000)
        self.sales_quantity_spinbox.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
        self.sales_quantity_spinbox.config(state="disabled")

        # Price
        price_label = ttk.Label(frame_widgets, text="Price:")
        price_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.sales_price_spinbox = ttk.Spinbox(frame_widgets, from_=0.0, to=10000.0, increment=50)
        self.sales_price_spinbox.grid(row=4, column=1, padx=5, pady=5, sticky="ew")
        self.sales_price_spinbox.config(state="disabled")

        # Buttons for product operations
        self.sales_cancel_button = ttk.Button(frame_widgets, text="Cancel", command=self.reset_product_cancel)
        self.sales_cancel_button.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")
        self.sales_cancel_button.config(state="disabled")
        
        self.sales_return_button = ttk.Button(frame_widgets, text="Return", command=self.return_product)
        self.sales_return_button.grid(row=5, column=1, padx=5, pady=5, sticky="nsew")
        self.sales_return_button.config(state="disabled")

        # Separator for better UI structure
        separator = ttk.Separator(frame_widgets)
        separator.grid(row=6, column=0, columnspan=2, padx=(20, 10), pady=10, sticky="ew")
        
        # Create a DateEntry widget start
        date_start_label = ttk.Label(frame_widgets, text="Start Date")
        date_start_label.grid(row=7, column=0, padx=5, pady=5, sticky="w")
        self.date_start = DateEntry(frame_widgets, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.date_start.grid(row=7, column=1, padx=5, pady=5, sticky="ew")
        
        # Create a DateEntry widget end
        date_end_label = ttk.Label(frame_widgets, text="End Date")
        date_end_label.grid(row=8, column=0, padx=5, pady=5, sticky="w")
        self.date_end = DateEntry(frame_widgets, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.date_end.grid(row=8, column=1, padx=5, pady=5, sticky="ew")
        
        # 
        calculate_sales_button = ttk.Button(frame_widgets, text="Total", command=self.calculate_sales)
        calculate_sales_button.grid(row=9, column=0, padx=5, pady=5, sticky="nsew")
        self.calculate_sales_entry = ttk.Entry(frame_widgets)
        self.calculate_sales_entry.grid(row=9, column=1, padx=5, pady=5, sticky="ew")
        self.calculate_sales_entry.config(state="disabled")

    def add_berevage_widgets(self, frame):
        """Creates and adds the widgets for the product management section."""
        frame_widgets = ttk.LabelFrame(frame, text="Berevage")
        frame_widgets.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        
        # Configure grid weights to make columns and rows responsive
        frame_widgets.grid_columnconfigure(1, weight=1)  # Make the second column (with entry fields) expandable
        for i in range(9):
            frame_widgets.grid_rowconfigure(i, weight=1)  # Make each row expandable
        
        # Product Name
        name_label = ttk.Label(frame_widgets, text="Name:")
        name_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.beverage_name_entry = ttk.Entry(frame_widgets)
        self.beverage_name_entry.bind("<FocusIn>", lambda e: self.clear_entry(self.product_name_entry, "Name"))
        self.beverage_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        # Category
        category_label = ttk.Label(frame_widgets, text="Category:")
        category_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.beverage_category_combobox = ttk.Combobox(frame_widgets, values=self.combo_list_beverageCategory, state="readonly")
        self.beverage_category_combobox.current(0)  # Default to the first category
        self.beverage_category_combobox.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        # Brand
        type_label = ttk.Label(frame_widgets, text="Brand:")
        type_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.beverage_brand_combobox = ttk.Combobox(frame_widgets, values=self.combo_list_beveragebrand, state="readonly")
        self.beverage_brand_combobox.current(0)
        self.beverage_brand_combobox.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        # Quantity
        quantity_label = ttk.Label(frame_widgets, text="Quantity:")
        quantity_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.beverage_quantity_spinbox = ttk.Spinbox(frame_widgets, from_=1, to=1000)
        self.beverage_quantity_spinbox.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        # Price
        price_label = ttk.Label(frame_widgets, text="Price:")
        price_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.beverage_price_spinbox = ttk.Spinbox(frame_widgets, from_=0.0, to=500.0, increment=10)
        self.beverage_price_spinbox.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

        # Buttons for product operations , command=self.insert_product
        self.beverage_save_button = ttk.Button(frame_widgets, text="Insert", command=self.Insert_beverage)
        self.beverage_save_button.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")
        # , command= lambda :self.update_product(treeview=self.treeview1)
        self.beverage_update_button = ttk.Button(frame_widgets, text="Update")
        self.beverage_update_button.grid(row=5, column=1, padx=5, pady=5, sticky="nsew")
        self.beverage_update_button.config(state="disabled")
        # , command=lambda: self.delete_product(treeview=self.treeview1)
        self.beverage_delete_button = ttk.Button(frame_widgets, text="Delete")
        self.beverage_delete_button.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")
        self.beverage_delete_button.config(state="disabled")
        # , command=self.reset_product
        self.beverage_cancel_button = ttk.Button(frame_widgets, text="Cancel", command=self.reset_beverage)
        self.beverage_cancel_button.grid(row=6, column=1, padx=5, pady=5, sticky="nsew")
        self.beverage_cancel_button.config(state="disabled")

        # Separator for better UI structure
        separator = ttk.Separator(frame_widgets)
        separator.grid(row=7, column=0, columnspan=2, padx=(20, 10), pady=10, sticky="ew")
        
        # Sales Button , command=self.open_sales_window
        self.beverage_Sales_button = ttk.Button(frame_widgets, text="Sale")
        self.beverage_Sales_button.grid(row=8, column=0, columnspan=2, padx=5, pady=10, sticky="nsew")
        self.beverage_Sales_button.config(state="disable")

    def create_treeview(self, treeview_attr, frame, path, columns):
        """Creates the treeview widget for displaying product data."""
        # Create a frame to hold the treeview and scrollbar
        tree_frame = ttk.Frame(frame)
        tree_frame.grid(row=1, column=1, padx=20, pady=10, sticky="nsew")
        
        # Configure grid weights for the tree_frame to make it expandable
        
        # Create a vertical scrollbar
        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.grid(row=0, column=1, sticky="ns")

        # Create the Treeview widget
        treeview = ttk.Treeview(tree_frame, show="headings", yscrollcommand=tree_scroll.set, columns=columns, height=1)
        
        # Configure Treeview columns and headings
        for col in columns:
            treeview.column(col, minwidth=100, anchor="center")
            treeview.heading(col, text=col, command=lambda _col=col: self.sort_treeview(_col, False, treeview=treeview))

        # Place Treeview in the grid and make it expand to fill available space
        treeview.grid(row=0, column=0, pady=10, sticky="nsew")

        # Configure the Scrollbar to control the Treeview's vertical scrolling
        tree_scroll.config(command=treeview.yview)

        # Bind selection event to Treeview
        treeview.bind("<<TreeviewSelect>>", lambda event: self.on_item_selected(event, treeview=treeview))
        
        # Store the Treeview object to be accessible from other parts of the class
        setattr(self, treeview_attr, treeview)

        # Load data into the Treeview
        self.load_data(treeview=treeview, path=path)
        
        # Configure grid weights for the treeview and scrollbar to allow responsiveness
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def page_title(self, frame, title):
        """Creates and displays the page title."""
        # Create a frame for the title to manage layout
        title_frame = ttk.Frame(frame)
        title_frame.grid(row=0, column=1, columnspan=2, padx=20, pady=10, sticky="ew")

        # Create a label with large text for the title
        title_label = ttk.Label(title_frame, text=title, font=("Helvetica", 24, "bold"))
        title_label.grid(row=0, column=0, padx=10, pady=10, sticky="n")
        title_label.config(anchor="center", foreground="#333333")

        # Adjust column weights to center the title
        title_frame.columnconfigure(0, weight=1)

    def create_search(self, frame, search_entry_name, path):
            
        """Creates the search bar for filtering products in the treeview."""
        search_frame = ttk.Frame(frame)
        search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="ew")

        search_label = ttk.Label(search_frame, text="Search:")
        search_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        # Create a single search entry attribute, regardless of the path
        search_entry = ttk.Entry(search_frame)
        search_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        # Store the search_entry object to be accessible from other parts of the class
        setattr(self, search_entry_name, search_entry)

        # Bind the appropriate treeview depending on the path
        if path == "products.xlsx":
            self.product_search_entry.bind("<KeyRelease>", lambda event: self.search_data(event, search_entry=search_entry, treeview=self.treeview1))
        else:
            self.sales_search_entry.bind("<KeyRelease>", lambda event: self.search_data(event, search_entry=search_entry, treeview=self.treeview2))

        
        # Configure grid weight to ensure proper resizing
        search_frame.columnconfigure(1, weight=1)

# GUI PART ------------------------------------------------------------------------------


# FUN PART ------------------------------------------------------------------------------

# TREEVIEW FUNCTIONALITY

    def sort_treeview(self, col, reverse, treeview):
        """Sorts the Treeview column when the heading is clicked."""
        try:
            data = [(treeview.set(k, col), k) for k in treeview.get_children('')]
            data.sort(reverse=reverse)
            for index, (val, k) in enumerate(data):
                treeview.move(k, '', index)
            treeview.heading(col, command=lambda: self.sort_treeview(col, not reverse, treeview=treeview))
        except Exception as e:
            print(f"Error sorting column {col}: {e}")
    
    def on_item_selected(self, event, treeview):
        selected_item = treeview.selection()  # Get the selected item
        if selected_item:
            if  treeview == self.treeview1:
                self.reset_product_flag = True
                # Retrieve the values of the selected item
                item_values = treeview.item(selected_item, "values")
                
                # Populate the input fields with the selected item's data
                self.product_name_entry.config(state="normal")
                self.product_name_entry.delete(0, "end")
                self.product_name_entry.insert(0, item_values[0])  # Name
                self.product_name_entry.config(state="disabled")
                
                self.category_combobox.set(item_values[1])  # Category
                
                if  item_values[1] == "Flash USB" or item_values[1] == "Memory Card":
                    self.memoryType_combobox.set(item_values[2])  # Type
                    self.memoryType_combobox.config(state="normal")
                    self.type_combobox.config(state="disable")
                else:
                    self.type_combobox.set(item_values[2])  # Type
                    self.type_combobox.config(state="normal")
                    self.memoryType_combobox.config(state="disable")
                
                self.quantity_spinbox.delete(0, "end")
                self.quantity_spinbox.insert(0, item_values[3])  # Quantity
                
                self.price_spinbox.delete(0, "end")
                self.price_spinbox.insert(0, item_values[4])  # Price
                
                # Disable the Insert button and enable the Update and Delete buttons
                self.insert_button.config(state="disabled")
                self.update_button.config(state="normal")
                self.delete_button.config(state="normal")
                self.cancel_button.config(state="normal")
                self.Sales_button.config(state="normal")
                
                # Store the selected item's ID for future updates
                self.selected_prodcut_item = item_values[0]
            elif treeview == self.treeview2:
                self.reset_sales_flag = True
                # Retrieve the values of the selected item
                item_values = treeview.item(selected_item, "values")
                self.sales_name_entry.config(state="normal")
                self.sales_name_entry.delete(0, "end")
                self.sales_name_entry.insert(0, item_values[1])  # Name
                self.sales_name_entry.config(state="disabled")
                
                self.sales_category_combobox.config(state="normal")
                self.sales_category_combobox.set(item_values[2])
                self.sales_category_combobox.config(state="disabled")
                
                self.sales_type_entry.config(state="normal")
                self.sales_type_entry.delete(0,  "end")
                self.sales_type_entry.insert(0, item_values[3])
                self.sales_type_entry.config(state="disabled")
                
                self.sales_quantity_spinbox.config(state="normal")
                self.sales_quantity_spinbox.delete(0, "end")
                self.sales_quantity_spinbox.insert(0, item_values[4])
                
                self.sales_price_spinbox.config(state="normal")
                self.sales_price_spinbox.delete(0, "end")
                self.sales_price_spinbox.insert(0, item_values[5])
                self.sales_price_spinbox.config(state="disabled")

                
                self.sales_return_button.config(state="normal")
                self.sales_cancel_button.config(state="normal")
                
                # Store the selected item's ID 
                self.selected_sales_item = [item_values[0],  item_values[1], item_values[4]]
            else:
                self.reset_beverage_flag = True
                # Retrieve the values of the selected item
                item_values = treeview.item(selected_item, "values")
                
                # Populate the input fields with the selected item's data
                self.beverage_name_entry.config(state="normal")
                self.beverage_name_entry.delete(0, "end")
                self.beverage_name_entry.insert(0, item_values[1])  # Name
                self.beverage_name_entry.config(state="disabled")

                self.beverage_category_combobox.set(item_values[2])  # Category
                
                self.beverage_brand_combobox.set(item_values[3])  # Type
                self.beverage_brand_combobox.config(state="normal")
                
                self.beverage_quantity_spinbox.delete(0, "end")
                self.beverage_quantity_spinbox.insert(0, item_values[4])  # Quantity
                
                self.beverage_price_spinbox.delete(0, "end")
                self.beverage_price_spinbox.insert(0, item_values[5])  # Price
                
                # Disable the Insert button and enable the Update and Delete buttons
                self.beverage_save_button.config(state="disabled")
                self.beverage_update_button.config(state="normal")
                self.beverage_delete_button.config(state="normal")
                self.beverage_cancel_button.config(state="normal")
                self.beverage_Sales_button.config(state="normal")
                
                self.selected_beverage_item = item_values[0]
                
    def load_data(self, treeview, path):
        """Loads data from the Excel file and inserts it into the treeview."""
        try:
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            
            if path == "products.xlsx":
                self.data_product = list(sheet.values)  # Store all data for searching
                all_data = self.data_product
            else:
                self.data_sales = list(sheet.values)  # Store all data for searching
                all_data = self.data_sales
            headers = all_data[0]
            treeview["columns"] = headers
            for col in headers:
                treeview.heading(col, text=col, command=lambda _col=col: self.sort_treeview(_col, False, treeview=treeview))
                treeview.column(col, width=100, anchor="center")

            treeview.delete(*treeview.get_children())  # Clear existing data

            for value_tuple in all_data[1:]:
                treeview.insert('', tk.END, values=value_tuple)
        except FileNotFoundError:
            # If the file doesn't exist, create it with headers
            self.create_excel_file(path)
        except Exception as e:
            print(f"Error loading data: {e}")

    def create_excel_file(self, path):
        """Creates a new Excel file with headers."""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = ["Name", "Category", "Type", "Quantity", "Price", "Date", "Time"]
            headers_sales = ["Id", "Name", "Category", "Type", "Quantity", "Price", "Date", "Time"]
            headers_beverage = ["Id", "Name", "Category", "Brand", "Quantity", "Price", "Date", "Time"]
            if  path == "products.xlsx":
                sheet.append(headers)
                self.data_product = [headers]
            elif path == "sales.xlsx":
                sheet.append(headers_sales)
                self.data_sales = [headers_sales]
            else:
                sheet.append(headers_beverage)
                self.data_sales = [headers_beverage]
            workbook.save(path)
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            
    def on_category_selected(self, event):
        selected_category = self.category_combobox.get()
        if selected_category == "Flash USB" or selected_category == "Memory Card":  # Specify the category you want to check
            self.memoryType_combobox.config(state="normal")  # Enable the type combobox
            self.type_combobox.config(state="disable")  # Disable the type combobox
        else:
            self.type_combobox.config(state="normal")  # normal the type combobox
            self.memoryType_combobox.config(state="disable")  # Enable the type combobox

    def clear_entry(self, entry, default_text):
        if entry.get() == default_text:
            entry.delete(0, "end")

# PRODUCT UPDATE FUNCTIONALITY
    
    def update_product(self, treeview):
        if not self.selected_prodcut_item:
            messagebox.showwarning("Select Item", "Please select an item to update.")
            return

        # Show confirmation dialog
        response = messagebox.askyesno("Confirm Update", "Are you sure you want to update this product?")
    
        if response:
            validated_data = self.validate_inputs()
            if not validated_data:
                return

            name, category, product_type, quantity, price = validated_data
            current_date = datetime.now().strftime("%Y-%m-%d")
            current_time = datetime.now().strftime("%H:%M:%S")
            
            # Prepare new data
            new_data = [name, category, product_type, quantity, price, current_date, current_time]
            
            # Update Treeview
            for item in treeview.get_children():
                if treeview.item(item, "values")[0] == self.selected_prodcut_item:
                    treeview.item(item, values=new_data)
                    break
            
            # Update Excel file
            path = "products.xlsx"
            try:
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active

                for row in sheet.iter_rows(values_only=False):
                    if row[0].value == name: 
                        row[1].value = category
                        row[2].value = product_type
                        row[3].value = quantity
                        row[4].value = price
                        row[5].value = current_date
                        row[6].value = current_time
                        break

                workbook.save(path)
                messagebox.showinfo("Success", "Product updated successfully!")
                self.reset_product()
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while updating the product: {e}")
  
# PRODUCT DELETE FUNCTIONALITY
    
    def delete_product(self, treeview):
        
        if not self.selected_prodcut_item:
            messagebox.showwarning("Select Item", "Please select an item to delete.")
            return
        
        # Show confirmation dialog
        response = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this product?")
        
        if response:
            # Remove from Treeview
            for item in treeview.get_children():
                if treeview.item(item, "values")[0] == self.selected_prodcut_item:
                    treeview.delete(item)
                    break

            # Remove from Excel file
            path = "products.xlsx"
            try:
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active

                for i, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                    if row[0].value == self.selected_prodcut_item:
                        sheet.delete_rows(i, 1)
                        break

                workbook.save(path)
                messagebox.showinfo("Success", "Product deleted successfully!")
                self.reset_product()
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while deleting the product: {e}")

# PRODUCT INSERTION FUNCTIONALITY

    def insert_product(self):
        validated_data = self.validate_inputs()
        if not validated_data:
            return
        
        name, category, product_type, quantity, price = validated_data
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")
        
        # unique_id = str(uuid.uuid4())
        row_values = [name, category, product_type, quantity, price, current_date, current_time]
        
        # Insert in a separate thread to keep UI responsive
        threading.Thread(target=self._insert_product, args=(row_values,)).start()
    
    def _insert_product(self, row_values):
        path = "products.xlsx"
        try:
            if not os.path.exists(path):
                self.create_excel_file(path)
            
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active

            # Check for duplicate product name
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0].lower() == row_values[0].lower():  # Assuming Name is unique
                    messagebox.showwarning("Product Exists", f"The product '{row_values[0]}' already exists.")
                    return

            # Append the new product
            sheet.append(row_values)
            workbook.save(path)
            
            # Insert into Treeview
            self.treeview1.insert('', tk.END, values=row_values)
            
            # Update the stored data
            self.data_beverage.append(row_values)
            
            # Clear the input fields
            self.reset_product()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while inserting the product: {e}")
    
    def validate_inputs(self):
        try:
            name = self.product_name_entry.get().strip()
            if not name:
                messagebox.showerror("Error", "Please enter a valid Product Name!")
                return None

            category = self.category_combobox.get()
            if category == "Flash USB" or category == "Memory Card":  # Specify the category you want to check
                product_type = self.memoryType_combobox.get()
            else:
                product_type = self.type_combobox.get()
            quantity = int(self.quantity_spinbox.get())
            price = float(self.price_spinbox.get())
            
            return name, category, product_type, quantity, price
        except ValueError:
            messagebox.showerror("Error", "Ensure all fields are entered correctly!")
            return None
        
# BEVERAGE INSERTION FUNCTIONALITY

    def Insert_beverage(self):
        validated_data = self.validate_beverage_input()
        if not validated_data:
            return
        
        id, name, category, brand, quantity, price = validated_data
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")
        row_values = [id, name, category, brand, quantity, price, current_date, current_time]
        
        # Insert in a separate thread to keep UI responsive
        threading.Thread(target=self._insert_beverage, args=(row_values,)).start()
    
    def _insert_beverage(self, row_values):
        path = "beverage.xlsx"
        try:
            if not os.path.exists(path):
                self.create_excel_file(path)
            
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            

            # Append the new product
            sheet.append(row_values)
            workbook.save(path)
            
            # Insert into Treeview
            self.treeview3.insert('', tk.END, values=row_values)
            
            # Update the stored data
            self.data_product.append(row_values)
            
            # Clear the input fields
            self.reset_beverage()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while inserting the product: {e}")
    
    def validate_beverage_input(self):
        try:
            name = self.beverage_name_entry.get().strip()
            if not name:
                messagebox.showerror("Error", "Please enter a valid Product Name!")
                return None

            category = self.beverage_category_combobox.get()
            brand = self.beverage_brand_combobox.get()
            quantity = int(self.beverage_quantity_spinbox.get())
            price = float(self.beverage_price_spinbox.get())
            
            # Generate a random UUID
            unique_id = uuid.uuid4()

            # Convert to string if needed
            unique_id_str = str(unique_id)
                
            return unique_id_str, name, category, brand, quantity, price
        except ValueError:
            messagebox.showerror("Error", "Ensure all fields are entered correctly!")
            return None

# RETURN PRODUCT FUNCTIONALITY
    
    def return_product(self):
        if not self.selected_sales_item:
            messagebox.showwarning("Select Item", "Please select an item to return.")
            return

        # Load the quantity to return
        try:
            self.return_quantity = int(self.sales_quantity_spinbox.get())
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid quantity.")
            return

        # Confirm return action
        if not messagebox.askyesno("Confirm Return", f"Are you sure you want to return {self.return_quantity} of '{self.selected_sales_item[1]}'?"):
            return

        self.return_sales_flag = True

        # Perform the return
        try:
                self._remove_from_sales_file("sales.xlsx")
                self._update_products_file("products.xlsx")
                messagebox.showinfo("Success", "Product return successful!")
                self.reset_sales()
                self.return_sales_flag = False

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while returning the product: {e}")

    def _remove_from_sales_file(self, sales_path):
        sales_workbook = openpyxl.load_workbook(sales_path)
        sales_sheet = sales_workbook.active
        

        # Find and update the row in sales.xlsx
        for i, row in enumerate(sales_sheet.iter_rows(values_only=False), start=1):
            if str(row[0].value) == str(self.selected_sales_item[0]):  # Assuming column 0 is the sales item ID
                if row[4].value is None or not isinstance(row[4].value, (int, float)):
                    raise ValueError("Invalid or missing quantity in the sales record.")
                if row[4].value > self.return_quantity:
                    row[4].value -= self.return_quantity
                elif row[4].value == self.return_quantity:
                    sales_sheet.delete_rows(i, 1)
                else:
                    raise ValueError("Return quantity exceeds available quantity.")
                break

        sales_workbook.save(sales_path)

    def _update_products_file(self, products_path):
        products_workbook = openpyxl.load_workbook(products_path)
        products_sheet = products_workbook.active
        
        print("_update_products_file")

        # Update the product quantity in products.xlsx
        for row in products_sheet.iter_rows(values_only=False):
            if row[0].value == self.selected_sales_item[1]:  # Assuming column 0 is the product ID
                row[3].value += self.return_quantity  # Assuming column 3 is the product quantity
                break
            
        

        products_workbook.save(products_path)


# SEARCH FUNCTIONALITY 
    
    def search_data(self, event, search_entry, treeview):
        if self.reset_product_flag:
            self.reset_product()
        search_term = search_entry.get().lower()
        treeview.delete(*treeview.get_children())
        
        if treeview == self.treeview1:
            all_data = self.data_product
        else:
            all_data = self.data_sales

        for value_tuple in all_data[1:]:
            if any(search_term in str(cell).lower() for cell in value_tuple):
                treeview.insert('', tk.END, values=value_tuple)
           
                
# RESET FUNCTIONALITY

    def reset_product(self):
        if self.selected_prodcut_item:
            for item in self.treeview1.get_children():
                if self.treeview1.item(item, "values")[0] == self.selected_prodcut_item:
                    self.treeview1.selection_remove(item)
                    break
        self.selected_prodcut_item = None
        
        # Reset on the product part 
        self.insert_button.config(state="normal")
        self.type_combobox.config(state="disabled")
        self.memoryType_combobox.config(state="disabled")
        self.update_button.config(state="disabled")
        self.delete_button.config(state="disabled")
        self.cancel_button.config(state="disabled")
        self.Sales_button.config(state="disabled")
        # Reset input fields
        self.product_name_entry.config(state="normal")
        self.product_name_entry.delete(0, "end")
        self.category_combobox.set(self.combo_list_ProductCategory[0])
        self.type_combobox.set(self.combo_list_ProductType[0])
        self.quantity_spinbox.delete(0, "end")
        self.price_spinbox.delete(0, "end")
        
        #  Reset on the sales part 
        self.reset_product_flag = False

    def reset_beverage(self):
        if self.selected_beverage_item:
            for item in self.treeview3.get_children():
                if self.treeview3.item(item, "values")[0] == self.selected_beverage_item:
                    self.treeview3.selection_remove(item)
                    break
        self.selected_beverage_item = None
        
        # Reset on the product part 
        self.beverage_save_button.config(state="normal")
        self.beverage_update_button.config(state="disabled")
        self.beverage_delete_button.config(state="disabled")
        self.beverage_cancel_button.config(state="disabled")
        self.beverage_Sales_button.config(state="disabled")
        
        # Reset input fields
        self.beverage_name_entry.config(state="normal")
        self.beverage_name_entry.delete(0, "end")
        self.beverage_category_combobox.set(self.combo_list_ProductCategory[0])
        self.beverage_brand_combobox.set(self.combo_list_ProductType[0])
        self.beverage_quantity_spinbox.delete(0, "end")
        self.beverage_price_spinbox.delete(0, "end")
        
        self.reset_beverage_flag = False

    def reset_product_cancel(self):
        if self.selected_sales_item:
            # Remove Selection item from Treeview2
            for item in self.treeview2.get_children():
                if self.treeview2.item(item, "values")[0] == self.selected_sales_item[0]:
                    self.treeview2.selection_remove(item)
                    break
            
            self.reset_product_entry()
            
    def reset_product_entry(self):
        # Reset entry fields
        self.sales_name_entry.config(state="normal")
        self.sales_name_entry.delete(0, "end")
        self.sales_name_entry.config(state="disabled")
        
        self.sales_category_combobox.set(self.combo_list_ProductCategory[0])
        self.sales_category_combobox.config(state="disabled")

        self.sales_type_entry.config(state="normal")
        self.sales_type_entry.delete(0, "end")
        self.sales_type_entry.config(state="disabled")

        self.sales_quantity_spinbox.delete(0, "end")
        self.sales_quantity_spinbox.config(state="disabled")

        self.sales_price_spinbox.config(state="normal")
        self.sales_price_spinbox.delete(0, "end")
        self.sales_price_spinbox.config(state="disabled")

        self.sales_return_button.config(state="disabled")
        self.sales_cancel_button.config(state="disabled")
              
    def reset_sales(self):
        if self.selected_sales_item:
            try:
                selected_quantity = int(self.selected_sales_item[2])
                if selected_quantity < self.return_quantity:
                    messagebox.showerror("Error", f"the return quantity is bigger than the sales quantity in the table please insert quantity less than {selected_quantity}.")
                    return 
            except (ValueError, IndexError):
                messagebox.showerror("Error", "Invalid quantity or item selection.")
                return 
            if self.return_sales_flag and selected_quantity > self.return_quantity:
                
                # Update ITEM FROM Treeview2
                self.update_treeview2()
            else:
                
                # Delete item from Treeview2
                for item in self.treeview2.get_children():
                    if self.treeview2.item(item, "values")[0] == self.selected_sales_item[0]:
                        self.treeview2.delete(item)
                        break
                
            # Update ITEM FROM Treeview1
            self.update_treeview1()
            
        self.selected_sales_item.clear()

        # Reset entry fields
        self.reset_product_entry()

        self.reset_sales_flag = False
        
# Update tables for the return fun ..     
    def update_treeview2(self):
        for item in self.treeview2.get_children():
            if self.treeview2.item(item, "values")[0] == self.selected_sales_item[0]:
                values = list(self.treeview2.item(item, "values"))
                values[4] = str(int(values[4]) - self.return_quantity)
                self.treeview2.item(item, values=values)
                self.treeview2.selection_remove(item)
                break

    def update_treeview1(self):
        for item in self.treeview1.get_children():
            if self.treeview1.item(item, "values")[0] == self.selected_sales_item[1]:
                values = list(self.treeview1.item(item, "values"))
                values[3] = str(int(values[3]) + self.return_quantity)
                self.treeview1.item(item, values=values)
                self.treeview1.selection_remove(item)
                break

    # Function to calculate total sales between two dates
    def calculate_sales(self):
        file_path = 'sales.xlsx'  
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        total_sales_amount = 0
        start_date_ = self.date_start.get_date().strftime('%Y-%m-%d')
        end_date_ = self.date_end.get_date().strftime('%Y-%m-%d')
        start_date = datetime.strptime(start_date_, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_, '%Y-%m-%d')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date_str = row[6]  # Assuming 'Date' is in the 7th column (index 6)
            if isinstance(date_str, str):
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')

                if start_date <= date_obj <= end_date:
                    quantity = float(row[4])  # Quantity in the 5th column (index 4)
                    price = float(row[5])     # Price in the 6th column (index 5)
                    total_sales_amount += quantity * price
        self.calculate_sales_entry.config(state="normal")
        self.calculate_sales_entry.delete(0, "end")
        self.calculate_sales_entry.insert(0, string=f"{total_sales_amount} DA")
        self.calculate_sales_entry.config(state="disabled")
                    
#FUN PART ------------------------------------------------------------------------------

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelCrafterApp(root)
    root.mainloop()